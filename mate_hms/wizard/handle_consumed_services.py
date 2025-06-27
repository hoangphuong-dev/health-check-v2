# -*- coding: utf-8 -*-
import logging

from odoo import models, api, fields, _
from collections import Counter
from odoo.exceptions import ValidationError
from datetime import timedelta

MATE_HMS_APPOINTMENT = 'mate_hms.appointment'

_logger = logging.getLogger(__name__)

try:
    from openpyxl import load_workbook
except ImportError:
    _logger.warning("Please install openpyxl to handle Excel files.")
    load_workbook = None


class MateHandleConsumedServices(models.TransientModel):
    _name = 'mate_hms.handle.consumed.services'
    _inherit = ['mate.excel.handler.base']
    _description = "Wizard to upload and handle consumed services from Excel file"

    appointment_id = fields.Many2one(MATE_HMS_APPOINTMENT, string='Appointment')
    excel_file = fields.Binary()
    excel_file_name = fields.Char()
    # - consumed_services_line_ids: Dùng để lưu những services được upload lên từ file excel.
    consumed_services_line_ids = fields.One2many('mate_hms.consumed.services.line', 'handle_consumed_services_id', string='Consumed Services Line', required=True)
    date = fields.Datetime(string='Date', store=True)
    date_to = fields.Datetime(string='Date To', store=True)

    @api.onchange('consumed_services_line_ids')
    def onchange_consumed_services_line_ids(self):
        """
        Hàm check khi sửa hoặc xóa consumed_services_line_ids thì sẽ check trùng lặp code
        và sau đó đánh duplicated lại cho các bản ghi
        :return:
        """
        if self.consumed_services_line_ids:
            codes = [line.code for line in self.consumed_services_line_ids]
            self.consumed_services_line_ids = self._handle_duplicate_code(self.consumed_services_line_ids, codes)

    @api.onchange('excel_file')
    def onchange_excel_file(self):
        if self.excel_file and load_workbook:
            self._validate_file_name(self.excel_file_name)
            header, data_rows = self._get_header_and_data(self.excel_file)
            if not data_rows:
                raise ValidationError(_("No data found in the uploaded file."))
            list_data, codes, earliest_time = self._process_excel_rows(header, data_rows)

            # Set appointment date if we found a earliest time
            if earliest_time:
                self.date = earliest_time
                self.date_to = earliest_time + timedelta(minutes=15)

            list_data = self._handle_duplicate_code(list_data, codes)
            self.consumed_services_line_ids = [(0, 0, line) for line in list_data]

    def _process_excel_rows(self, header, data_rows):
        list_data = []
        codes = []
        earliest_time = None

        for row_index, row in enumerate(data_rows, start=2):
            self._validate_row_not_empty(row_index, row)

            row_data = dict(zip(header, row))
            code = row_data.get('Mã')
            product_name = row_data.get('Nội dung')
            unit_price = row_data.get('Đơn giá')
            quantity = row_data.get('Số lượng')
            request_time_str = row_data.get('Thời gian yêu cầu')

            # Parse datetime if available
            request_time = self._parse_request_time(request_time_str)

            if request_time and (earliest_time is None or request_time < earliest_time):
                earliest_time = request_time

            no_error = self._validate_fields_excel(row_index, [code, product_name, unit_price, quantity])

            if no_error:
                codes.append(code)
                list_data.append({
                    'code': code,
                    'name': product_name,
                    'unit_price': unit_price or 0,
                    'quantity': quantity or 1,
                    'duplicated': False
                })

        return list_data, codes, earliest_time

    def _validate_fields_excel(self, index, data):
        """
        Hàm validate các trường dữ liệu trong file excel
        :param index:
        :param data:
        :return:
        """
        error_type = None

        code, product_name, unit_price, quantity = data

        if not code or not isinstance(code, (int, str)):
            error_type = _('Code')

        elif not product_name or not isinstance(product_name, str):
            error_type = _('Content')

        elif not unit_price or not isinstance(unit_price, (int, float)):
            error_type = _('Unit Price')

        elif not quantity or not isinstance(quantity, (int, float)):
            error_type = _('Quantity')

        if error_type:
            self._notify_error(str(index), error_type)

        return True

    def _check_duplicate_code(self, codes):
        """
        Hàm kiểm tra danh sách mã có trùng lặp không.
        Trả về:
        - [danh sách mã trùng] nếu có trùng
        - [] nếu không trùng
        """
        counter = Counter(codes)
        duplicates = [code for code, count in counter.items() if count > 1]
        return duplicates

    def _handle_duplicate_code(self, data, codes):
        """
        Hàm xử lý tíck những bản ghi duplicate code trong file excel
        :param data:
        :param codes:
        :return:
        """
        # Check trùng lặp code
        duplicate_codes = self._check_duplicate_code(codes)
        list_data = data
        for item in list_data:
            if item['code'] in duplicate_codes:
                item['duplicated'] = True
            else:
                item['duplicated'] = False

        return list_data

    def _get_appointment(self):
        return self.appointment_id

    def _generate_consumable_line(self, appointment, product, item):
        return {
            'appointment_id': appointment.id,
            'product_id': product.id,
            'price_unit': item.unit_price,
            'product_uom_id': product.uom_id.id,
            'qty': item.quantity,
        }

    def save_services_appointments(self):
        """
        Hàm xử lý logic khi nhấn nút Lưu
        :return:
        """
        if not self.excel_file:
            raise ValidationError(_("Please upload excel file!"))

        if not self.consumed_services_line_ids:
            raise ValidationError(_("There is no consumed services in the uploaded excel file!"))

        appointment = self._get_appointment()

        product_model = self.env['product.product']
        consumable_line_model = self.env['mate_hms.consumable.line']
        total_price = appointment.amount_total or 0

        # Lấy danh sách code trong consumed_services_line_ids
        codes = [item.code for item in self.consumed_services_line_ids]

        # Tìm sản phẩm trong database theo danh sách code
        products = product_model.search([('default_code', 'in', codes)])

        # Tạo 1 dictionary dạng {code: product}
        product_dict = {prod.default_code: prod for prod in products}

        for item in self.consumed_services_line_ids:
            # Xem sản phẩm đã tồn tại trong dictionary chưa
            product = product_dict.get(item.code)

            # Nếu không tồn tại, thì tạo sản phẩm mới
            if not product:

                product = product_model.create({
                    'name': item.name,
                    'default_code': item.code,
                    'standard_price': item.unit_price,
                    'list_price': item.unit_price,
                    'type': 'service',
                })
                # Thêm vào dictionary để tái sử dụng sau này
                product_dict[item.code] = product

            # Tính tổng giá của tất cả dịch vụ mà bệnh nhân sử dụng
            total_price += item.unit_price * item.quantity

            # Tạo consumable services và gắn vào appointment hiện tại
            consumable_line_model.create(self._generate_consumable_line(appointment, product, item))
        appointment.write({'amount_total': total_price})

        return appointment

    def download_template_file(self):
        """Download the template file"""
        return {
            'type': 'ir.actions.act_url',
            'url': '/mate_hms/static/description/sample_service.xlsx',
            'target': 'self',
        }
