from odoo import models, _
from odoo.exceptions import ValidationError
import base64
from io import BytesIO
from openpyxl import load_workbook
from zoneinfo import ZoneInfo
import logging

_logger = logging.getLogger(__name__)


class MateExcelHandlerBase(models.AbstractModel):
    _name = 'mate.excel.handler.base'
    _description = 'Base Excel handler for uploading and validating'

    def _validate_file_name(self, filename):
        if not filename.endswith(('.xls', '.xlsx')):
            raise ValidationError(_("Please upload files in .xls, .xlsx format!"))

    def _get_header_and_data(self, excel_file):
        file_content = base64.b64decode(excel_file)
        file_io = BytesIO(file_content)

        workbook = load_workbook(filename=file_io, read_only=True)
        sheet = workbook.active

        rows = list(sheet.iter_rows(values_only=True))
        if not rows or len(rows) < 2:
            raise ValidationError(_("The uploaded excel file has no data."))
        empty_row = 0
        for row in rows[1:]:
            if all(cell is None for cell in row):
                empty_row += 1
        if empty_row == len(rows[1:]):
            return rows[0], None
        return rows[0], rows[1:]

    def _validate_row_not_empty(self, row_index, row):
        if not any(row):
            raise ValidationError(_("Line %s has no data. Please check the file again!") % row_index)

    def _notify_error(self, line, error_type):
        raise ValidationError(_("Data in row %s of column %s is incorrect type or missing") % (line, error_type))

    def _parse_request_time(self, time_str):
        if not time_str:
            return False
        return time_str.replace(tzinfo=ZoneInfo("Asia/Ho_Chi_Minh")).astimezone(ZoneInfo("UTC")).replace(tzinfo=None)
